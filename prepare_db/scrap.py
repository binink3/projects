import time
import os
import os.path
import shutil
import xlrd
from openpyxl.workbook import Workbook
from openpyxl.reader.excel import load_workbook, InvalidFileException

<<<<<<< HEAD

=======
>>>>>>> 4bd07795997dc8f2a8f215cbba7b826e05a42d01
from datetime import datetime
from selenium import webdriver
from selenium.common.exceptions import TimeoutException
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.common.by import By
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.common.keys import Keys

<<<<<<< HEAD

now = datetime.now()
date = now.strftime("%Y%m%d")

download_path = 'C:/Users/lucke/Downloads'
file1 = '펀드매니저 상세정보_운용펀드(기준일)_'+ date + '.xls'
file2 = '펀드매니저 상세정보_운용펀드(과거3년)_'+ date + '.xls'
file_info = [file1, file2]

def dealing_alert(browser):
    try:
        WebDriverWait(browser, 3).until(EC.alert_is_present(), 'Timed out waiting for PA creation ' + 'confirmation popup to appear.')
=======
now = datetime.now()
date = now.strftime("%Y%m%d")

download_path = '/Users/giubinkang/Downloads'
file1 = '펀드매니저 상세정보_운용펀드(기준일)_' + date + '.xls'
file2 = '펀드매니저 상세정보_운용펀드(과거3년)_' + date + '.xls'
file_info = [file1, file2]


def dealing_alert(browser):
    try:
        WebDriverWait(browser, 3).until(EC.alert_is_present(),
                                        'Timed out waiting for PA creation ' + 'confirmation popup to appear.')
>>>>>>> 4bd07795997dc8f2a8f215cbba7b826e05a42d01
        alert = browser.switch_to.alert
        alert.accept()
        print("alert accepted")
    except TimeoutException:
        print("no alert")


def move_files(dir_path, file_prefix, filename):
    if os.path.exists(download_path + '/' + filename):
        shutil.move(download_path + '/' + filename, dir_path + '/' + file_prefix + filename)
        print(filename + " moved to " + dir_path)
        return dir_path + '/' + file_prefix + filename
    else:
        print(filename + " does not exist.")
        return "failure"


def convert_xls_to_xlsx(filename):
    # first open using xlrd
    book = xlrd.open_workbook(filename)
    xlsx_filename = filename[:filename.rfind('.')]
    index = 0
    nrows, ncols = 0, 0
    while nrows * ncols == 0:
        sheet = book.sheet_by_index(index)
        nrows = sheet.nrows
        ncols = sheet.ncols
        index += 1

    # prepare a xlsx sheet
    book1 = Workbook()
    sheet1 = book1.active

    for row in range(0, nrows):
        for col in range(0, ncols):
            sheet1.cell(row=row + 1, column=col + 1).value = sheet.cell_value(row, col)

    book1.save(filename='{}.xlsx'.format(xlsx_filename))
    return book1


def download_excel(driver, els):
    i = 0
    cnt = 0
    for el in els:
        if cnt >= 10:
            return
        if (i % 2 == 0):
            cnt += 1
            el.click()
            time.sleep(3)
            popup = driver.find_element_by_tag_name('iframe')
            driver.switch_to.frame(popup)

            time.sleep(3)
            name = driver.find_element_by_id('fundMgr').text
            comp = driver.find_element_by_id('compNm').text
            print(name, comp)
            file_prefix = comp + '_' + name + '_'
            dir_path = os.getcwd()

            driver.find_element_by_id('image111').click()

            # # li#tabControl1_tab_dtlTtabs2
            li = driver.find_element_by_id('tabControl1_tab_dtlTtabs2')
            btn = li.find_element_by_tag_name('a')
            btn.click()
            time.sleep(3)

            driver.find_element_by_id('image115').click()
            dealing_alert(driver)

            for file in file_info:
                filename = move_files(dir_path, file_prefix, file)
                if filename == 'failure':
                    continue
                convert_xls_to_xlsx(filename)

            driver.switch_to.default_content()
            close_btn = driver.find_element_by_id('textbox1987456321')
            close_btn.click()
        i += 1
<<<<<<< HEAD
=======


>>>>>>> 4bd07795997dc8f2a8f215cbba7b826e05a42d01
'''
1. 파일 다운로드
    1. 페이지 접속
    2. 목록 스캔
    3. 매 페이지 누르면서
        1. 파일 두 개 다운로드
        2. 펀드매니저_회사명 폴더 만든 다음에 다운 받은 파일 두 개를 이동
<<<<<<< HEAD
        
=======

>>>>>>> 4bd07795997dc8f2a8f215cbba7b826e05a42d01
2. 파일 DB에 저장
    1. 펀드매니저_회사명 폴더마다
        1. 파일 읽기
        2. DB 저장
'''

'''
driver
    - session_id {str}
    - current_url {str}
    - page_source {str}
    - title {str}
    - name {str} : browser

element
    - 
'''
<<<<<<< HEAD
driver = webdriver.Chrome('C:/Users/lucke/Downloads/chromedriver_win32/chromedriver.exe')
=======
driver = webdriver.Chrome('/Users/giubinkang/Downloads/chromedriver')
>>>>>>> 4bd07795997dc8f2a8f215cbba7b826e05a42d01

url = 'http://dis.kofia.or.kr/websquare/index.jsp?w2xPath=/wq/fundMgr/DISFundMgrSrch.xml&divisionId=MDIS03001002000000&serviceId=SDIS03001002000'
driver.get(url)

<<<<<<< HEAD

=======
>>>>>>> 4bd07795997dc8f2a8f215cbba7b826e05a42d01
element = WebDriverWait(driver, 300).until(EC.element_to_be_clickable((By.ID, 'btnSearImg')))
driver.find_element_by_id('btnSearImg').click()

element = WebDriverWait(driver, 300).until(EC.element_to_be_clickable((By.CLASS_NAME, 'w2grid_image')))

els = driver.find_elements_by_class_name('w2grid_image')

<<<<<<< HEAD


=======
>>>>>>> 4bd07795997dc8f2a8f215cbba7b826e05a42d01
try:
    for down in range(70):
        download_excel(driver, els)
        driver.switch_to.default_content()
        nth = driver.find_elements_by_class_name('gridBodyDefault')[1]
        nth.click()
        nth.send_keys(Keys.PAGE_DOWN)
        els = driver.find_elements_by_class_name('w2grid_image')
finally:
    driver.quit()

<<<<<<< HEAD

=======
>>>>>>> 4bd07795997dc8f2a8f215cbba7b826e05a42d01
# btnSearImg 클릭 img.w2grid_image
# 타겟 URL을 읽어서 HTML를 받아오고,
# headers = {'User-Agent' : 'Mozilla/5.0 (Windows NT 10.0; Win64; x64)AppleWebKit/537.36 (KHTML, like Gecko) Chrome/73.0.3683.86 Safari/537.36'}
# data = requests.get(url=url,headers=headers)
#
# # HTML을 BeautifulSoup이라는 라이브러리를 활용해 검색하기 용이한 상태로 만듦
# # soup이라는 변수에 "파싱 용이해진 html"이 담긴 상태가 됨
# # 이제 코딩을 통해 필요한 부분을 추출하면 된다.
# soup = BeautifulSoup(data.text, 'html.parser')
<<<<<<< HEAD
=======



>>>>>>> 4bd07795997dc8f2a8f215cbba7b826e05a42d01
