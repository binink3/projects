import os
from openpyxl import load_workbook

from pymongo import MongoClient           # pymongo를 임포트 하기(패키지 인스톨 먼저 해야겠죠?)
client = MongoClient('localhost', 27017)  # mongoDB는 27017 포트로 돌아갑니다.
db = client.publicfund                    # 'publicfund'라는 이름의 db를 만듭니다.

#Excel 파일의 제목을 List로 받아내기
path = os.getcwd()
file_list = os.listdir(path)
file_list_xlsx = [file for file in file_list
                if file.endswith(".xlsx")]

for name in file_list_xlsx:
    file_name = path+'/'+name
    work_book = load_workbook(file_name)
    work_sheet = work_book['sheet']

    company_name = name.split('_')[0]
    manager_name = name.split('_')[1]

    if file_name.find('과거3년') != -1:
        for row in work_sheet.iter_rows(min_row=3, max_row=10000):
            if row[0].value is None:
                break
            elif '-' not in row[0].value:
                fund_id_past = row[0].value
                fund_name_past = row[2].value
                fund_start_past = row[6].value
                fund_end_past = row[7].value
                fund_yield_past = row[17].value

                doc = {
                    'company_name' : company_name+'자산운용',
                    'manager_name' : manager_name,
                    'fund_id' : fund_id_past,
                    'fund_name' : fund_name_past,
                    'start_date' : fund_start_past,
                    'end_date' : fund_end_past,
                    'fund_yield' : fund_yield_past+'%',
                    'Status' : 'Past'
                }

                print (doc)
                #db.funds.insert_one(doc)


    elif file_name.find('기준일') != -1:
        for row in work_sheet.iter_rows(min_row=3, max_row=10000):
            if row[0].value is None:
                break
            elif '-' not in row[0].value:
                fund_id = row[0].value
                fund_name = row[1].value
                fund_start = row[5].value
                fund_yield = row[14].value

                doc2 = {
                    'company_name' : company_name+'자산운용',
                    'manager_name' : manager_name,
                    'fund_id' : fund_id,
                    'fund_name' : fund_name,
                    'start_date' : fund_start,
                    'end_date' : '-',
                    'fund_yield' : fund_yield+'%',
                    'Status' : 'Active'
                }

                print(doc2) #사실은 doc2를 doc으로 하더라도 두번째 elif구문내에서 돌아가는 변수이기 때문에, 위의 doc 변수와 겹칠 일이 없다.
                #db.funds.insert_one(doc2)